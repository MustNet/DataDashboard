import pandas as pd
import streamlit as st
import duckdb
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
import requests
from openpyxl.styles import PatternFill
from io import BytesIO

st.set_page_config(layout="wide")
st.title("Dashboard Logistics Data")

# URLs zu den Dateien im GitHub-Repository
url_dashboard1 = "https://raw.githubusercontent.com/MustNet/DataDashboard/main/Auftragsübersicht.xlsx"
url_dashboard2 = "https://raw.githubusercontent.com/MustNet/DataDashboard/main/Fahrposition.xlsx"
url_dashboard3 = "https://raw.githubusercontent.com/MustNet/DataDashboard/main/Transporte.xlsx"
url_dashboard4 = "https://raw.githubusercontent.com/MustNet/DataDashboard/main/Speditionspreise.xlsx"

# Funktion zum Herunterladen einer Datei von einer URL
def download_file_from_url(url):
    response = requests.get(url)
    if response.status_code == 200:
        return BytesIO(response.content)
    else:
        st.error(f"Fehler beim Herunterladen der Datei von {url}. Statuscode: {response.status_code}")
        return None

# Mapping der Statuscodes zu Statusmeldungen
zustand_mapping = {
    10: 'An Lvs Übertragen',
    15: 'Freigegeben',
    20: 'In Arbeit',
    40: 'Gestoppt',
    60: 'Auftrag abgeschlossen',
    65: 'Auftrag bereit für Verladung',
    70: 'Auftrag an ERP übertragen',
    80: 'Auftrag verladen'
}

# Definierte Farben für die Zustände
farben_mapping = {
    'In Arbeit': 'yellow',
    'Gestoppt': 'red',
    'Freigegeben': 'orange',
    'Auftrag verladen': 'green',
    'Auftrag abgeschlossen': 'blue',
    'An Lvs Übertragen': 'brown'
}

# Farben für die Jahre im Liniendiagramm
jahre_farben_mapping = {
    '2022': 'blue',
    '2023': 'green',
    '2024': 'red'
}

@st.cache_data
def load_data(file):
    # Verwende "converters", um die Spalten "Kd.-Nr.", "Auftrags-Nr." und "WWS-AuftragsNr." als String zu laden
    converters = {
        "Kd.-Nr.": lambda x: str(x).zfill(5),
        "Auftrags-Nr.": lambda x: str(x).zfill(5),
        "WWS-AuftragsNr.": lambda x: str(x).zfill(5)
    }
    
    # Excel-Datei mit den angegebenen Konvertierungen laden
    data = pd.read_excel(file, converters=converters)
    data = data.drop(index=0).reset_index(drop=True)
    data['Liefer-Dat.'] = pd.to_datetime(data['Liefer-Dat.'], errors='coerce')
    data['Zustand'] = data['Zustand'].map(zustand_mapping)
    data['Zustand'].fillna('Unbekannt', inplace=True)
    data['Jahr'] = data['Liefer-Dat.'].dt.year.astype(str)  # Jahr als String
    data['Monat'] = data['Liefer-Dat.'].dt.strftime('%b')  # Kürzel für den Monat (Jan, Feb, etc.)
    data['Monat_Zahl'] = data['Liefer-Dat.'].dt.month  # Monat als Zahl für die Sortierung
    return data

# Tabs für verschiedene Dashboards
tab1, tab2, tab3, tab4= st.tabs(["Dashboard 1", "Dashboard 2","Dashboard 3", "Dashboard 4"])

with tab1:
    st.subheader("Dashboard 1 - Auftragsübersicht_xlsx")

    # Datei automatisch von URL herunterladen
    file_dashboard1 = download_file_from_url(url_dashboard1)

    if file_dashboard1 is not None:
        # Checkbox zur Bestätigung durch den Benutzer
        if st.checkbox("Bestätigen Sie die Datei für Dashboard 1"):
            df = load_data(file_dashboard1)
            st.write(df.head())  # Zeige einen kurzen Überblick über die Daten

            # Filter für das Liniendiagramm (keine Zustandsfilterung)
            with st.sidebar:
                jahr_auswahl = st.selectbox("Wähle das Jahr", options=df["Jahr"].unique())
                monate_dict = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 
                               'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
                vorhandene_monate = sorted(df["Monat"].unique(), key=lambda x: monate_dict[x])
                monat_auswahl = st.selectbox("Wähle den Monat", options=vorhandene_monate)

            # 1. Gestapeltes Balkendiagramm: Anzahl der Zustände pro Tag im Monat
            df_balken = df[(df['Jahr'] == jahr_auswahl) & (df['Monat'] == monat_auswahl)]
            df_balken_grouped = df_balken.groupby(['Zustand', 'Liefer-Dat.']).size().reset_index(name='Anzahl')

            # 2. Gestapeltes Balkendiagramm: Anzahl der Zustände pro Monat im Jahr
            df_balken_jahr = df[df['Jahr'] == jahr_auswahl]
            df_balken_jahr_grouped = df_balken_jahr.groupby(['Zustand', 'Monat', 'Monat_Zahl']).size().reset_index(name='Anzahl')

            # Sortiere nach Monat_Zahl, um sicherzustellen, dass die Monate korrekt angeordnet sind
            df_balken_jahr_grouped = df_balken_jahr_grouped.sort_values('Monat_Zahl')

            col1, col2 = st.columns(2)

            with col1:
                st.subheader(f"Anzahl der Zustände pro Tag im {monat_auswahl} {jahr_auswahl}")
                fig_balken = px.bar(
                    df_balken_grouped,
                    x='Liefer-Dat.',
                    y='Anzahl',
                    color='Zustand',
                    color_discrete_map=farben_mapping,
                    labels={'Liefer-Dat.': 'Datum', 'Anzahl': 'Anzahl der Aufträge', 'Zustand': 'Zustand'},
                    title=f"Anzahl der Zustände pro Tag im {monat_auswahl} {jahr_auswahl}"
                )
                fig_balken.update_layout(xaxis_tickangle=-45, width=800, height=500)
                st.plotly_chart(fig_balken)

            with col2:
                st.subheader(f"Anzahl der Zustände pro Monat im Jahr {jahr_auswahl}")
                fig_balken_jahr = px.bar(
                    df_balken_jahr_grouped,
                    x='Monat',
                    y='Anzahl',
                    color='Zustand',
                    color_discrete_map=farben_mapping,
                    labels={'Monat': 'Monat', 'Anzahl': 'Anzahl der Aufträge', 'Zustand': 'Zustand'},
                    title=f"Anzahl der Zustände pro Monat im Jahr {jahr_auswahl}",
                    category_orders={'Monat': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']}
                )
                fig_balken_jahr.update_layout(width=800, height=500)
                st.plotly_chart(fig_balken_jahr)

            # Setze das Liniendiagramm neben das zweite Diagramm
            col3, col4 = st.columns(2)

            with col3:
                st.subheader(f"Anzahl der Aufträge pro Monat über alle Jahre (Gesamtanzahl)")
                query_3 = f"""
                    SELECT 
                        strftime('%m', "Liefer-Dat.") AS Monat_Zahl,
                        strftime('%b', "Liefer-Dat.") AS Monat,
                        Jahr,
                        COUNT(*) AS Anzahl_Aufträge
                    FROM df
                    GROUP BY Jahr, strftime('%b', "Liefer-Dat."), strftime('%m', "Liefer-Dat.")
                    ORDER BY strftime('%m', "Liefer-Dat.");
                """
                jahre_daten = conn.execute(query_3).df()
                jahre_daten['Monat_Zahl'] = jahre_daten['Monat_Zahl'].astype(int)

                monate = pd.DataFrame({
                    'Monat_Zahl': range(1, 13),
                    'Monat': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                })
                jahre_daten = jahre_daten.merge(monate, on='Monat_Zahl', how='right').fillna({'Anzahl_Aufträge': 0})

                jahre_daten['Monat'] = jahre_daten['Monat_x']
                jahre_daten = jahre_daten.drop(columns=['Monat_x', 'Monat_y'])
                jahre_daten = jahre_daten.sort_values('Monat_Zahl')

                fig_jahre = px.line(
                    jahre_daten, 
                    x='Monat', 
                    y='Anzahl_Aufträge', 
                    color='Jahr',
                    color_discrete_map=jahre_farben_mapping,
                    labels={'Monat': 'Monat', 'Anzahl_Aufträge': 'Anzahl der Aufträge', 'Jahr': 'Jahr'},
                    title="Anzahl der Aufträge pro Monat über alle Jahre (Gesamtanzahl)"
                )
                fig_jahre.update_layout(width=800, height=500)
                st.plotly_chart(fig_jahre)

            with col4:
                st.subheader(f"Aufträge nach Zuständen im {jahr_auswahl}")
                df_pie = df[df['Jahr'] == jahr_auswahl]
                df_pie_grouped = df_pie.groupby('Zustand').size().reset_index(name='Anzahl')

                fig_pie = px.pie(
                    df_pie_grouped,
                    names='Zustand',
                    values='Anzahl',
                    color='Zustand',
                    color_discrete_map=farben_mapping,
                    title="Anteil der Aufträge pro Zustand"
                )
                fig_pie.update_layout(width=800, height=500)
                st.plotly_chart(fig_pie)

            # Füge den Data Previewer wieder ein (optional)
            with st.expander("Data Preview"):
                st.dataframe(df)

    else:
        st.error("Die Datei für Dashboard 1 konnte nicht heruntergeladen werden.")


with tab2:
    st.subheader("Dashboard 2 - Fahrpositionen_xlsx")

    # Datei automatisch von URL herunterladen
    file_dashboard2 = download_file_from_url(url_dashboard2)

    if file_dashboard2 is not None:
        # Checkbox zur Bestätigung durch den Benutzer
        if st.checkbox("Bestätigen Sie die Datei für Dashboard 2"):
            df2 = load_data(file_dashboard2)
            st.write(df2.head())  # Zeige einen kurzen Überblick über die Daten
            # Restliche Visualisierungen hier hinzufügen...
    else:
        st.error("Die Datei für Dashboard 2 konnte nicht heruntergeladen werden.")

    # Lade die Daten für Dashboard 2 (nur die Datei hochladen und keine weiteren Veränderungen vornehmen)
    @st.cache_data
    def load_data_tab2(file):
        # Verwende "converters", um sicherzustellen, dass die Pers.-Nr. als String eingelesen wird
        converters = {
            "Pers.-Nr.": lambda x: str(x).zfill(4),  # Konvertiere die Pers.-Nr. in einen 4-stelligen String
            "Auftrags-Nr.": lambda x: str(x).zfill(5),        
        }

        # Excel-Datei mit den angegebenen Konvertierungen laden
        data = pd.read_excel(file, converters=converters)
        
        # Entferne führende und nachfolgende Leerzeichen von allen Spaltennamen
        data.columns = data.columns.str.strip()
        
        # Konvertiere "Ende Datum" in Datumsformat
        data['Ende Datum'] = pd.to_datetime(data['Ende Datum'], errors='coerce')
        
        # Konvertiere "Beginn Zeit" und "Ende Zeit" in Zeitformat für die Berechnung der Zeitdauer
        data['Beginn Zeit'] = pd.to_datetime(data['Beginn Zeit'], errors='coerce')
        data['Ende Zeit'] = pd.to_datetime(data['Ende Zeit'], errors='coerce')

        return data

    # Verwende die Funktion, um die Datei für Tab 2 zu laden
    df2 = load_data_tab2(uploaded_file_2)

    # Berechnung der Anzahl der "Gesamtpicks" und des "Gesamtgewichts" pro Personalnummer
    df2_grouped = df2.groupby('Pers.-Nr.').agg({
        'Anzahl Picks': 'sum',  # Summe der Picks pro Personalnummer
        'Gewicht': 'sum'  # Summe des Gewichts pro Personalnummer
    }).reset_index()

    # Neue Spalte mit dem Format "Personal + Personalnummer" hinzufügen
    df2_grouped['Personal'] = df2_grouped['Pers.-Nr.'].apply(lambda x: f"Personal {x}")

    # Setze drei Diagramme/Metriken nebeneinander
    col1, col2, col3 = st.columns(3)

    with col1:
        # Erstelle ein Balkendiagramm, das beide Metriken zeigt und die neue Personal-Spalte nutzt
        fig_balken = px.bar(
            df2_grouped,
            x='Personal',  # Nutze die neue Spalte 'Personal'
            y=['Gewicht', 'Anzahl Picks'],  # Zeige beide Metriken nebeneinander
            labels={'variable': 'Metrik', 'value': 'Wert', 'Personal': 'Personal'},
            title="Vergleich von Gesamtgewicht und Anzahl der Picks pro Personal",
            barmode='group'  # Nebeneinanderliegende Balken
        )

        # Größe des Diagramms anpassen
        fig_balken.update_layout(width=800, height=500)

        # Zeige das Balkendiagramm an
        st.plotly_chart(fig_balken)

    with col2:
        # Extrahiere das Jahr aus der Spalte "Ende Datum"
        df2['Jahr'] = df2['Ende Datum'].dt.year

        # Berechnung des Gesamtgewichts pro Jahr
        df2_jahr_grouped = df2.groupby('Jahr').agg({
            'Gewicht': 'sum'  # Summe des Gewichts pro Jahr
        }).reset_index()

        # Erstelle ein Kreisdiagramm für das Gesamtgewicht pro Jahr
        fig_pie = px.pie(
            df2_jahr_grouped,
            names='Jahr',
            values='Gewicht',
            title="Gesamtgewicht pro Jahr"
        )

        # Größe des Kreisdiagramms anpassen
        fig_pie.update_layout(width=800, height=500)

        # Zeige das Kreisdiagramm an
        st.plotly_chart(fig_pie)

    # Berechne die Gesamtanzahl der Aufträge
    gesamt_auftraege = df2['Auftrags-Nr.'].nunique()

    # Berechnung des Gesamtgewichts und des Durchschnittsgewichts
    gesamt_gewicht = df2['Gewicht'].sum()
    durchschnitt_gewicht = gesamt_gewicht / gesamt_auftraege if gesamt_auftraege > 0 else 0

    # Berechnung der Durchschnittsdauer für die Aufträge
    df2['Dauer'] = (df2['Ende Zeit'] - df2['Beginn Zeit']).dt.total_seconds() / 60  # Dauer in Minuten
    gesamt_dauer = df2['Dauer'].sum()
    durchschnitt_dauer = gesamt_dauer / gesamt_auftraege if gesamt_auftraege > 0 else 0

    with col3:
        st.subheader("Wichtige Kennzahlen")
        
        # Verwende st.columns, um die Metriken nebeneinander zu platzieren
        metric_col1, metric_col2, metric_col3 = st.columns(3)

        # Gesamtanzahl der kommissionierten Aufträge
        metric_col1.metric(label="Gesamtanzahl der Aufträge", value=f"{gesamt_auftraege}")

        # Durchschnittsgewicht pro Auftrag
        metric_col2.metric(label="Durchschnittsgewicht pro Auftrag", value=f"{durchschnitt_gewicht:.2f} kg")

        # Durchschnittsdauer pro Auftrag
        metric_col3.metric(label="Durchschnittszeit pro Auftrag", value=f"{durchschnitt_dauer:.2f} Minuten")

        # Liniendiagramm zur Entwicklung des Gewichts über die Jahre hinweg
        # Extrahiere Jahr und Monat aus "Ende Datum"
        df2['Monat'] = df2['Ende Datum'].dt.strftime('%b')
        df2['Monat_Zahl'] = df2['Ende Datum'].dt.month

        # Berechne das Gewicht pro Monat und Jahr in Tonnen (statt Kilogramm)
        df2_monate_grouped = df2.groupby(['Jahr', 'Monat', 'Monat_Zahl']).agg({
            'Gewicht': lambda x: x.sum() / 1000  # Summe des Gewichts in Tonnen
        }).reset_index()

        # Berechne das Gewicht pro Monat und Jahr
        df2_monate_grouped = df2.groupby(['Jahr', 'Monat', 'Monat_Zahl']).agg({
            'Gewicht': 'sum'
        }).reset_index()

        # Sortiere die Monate korrekt (Jan bis Dez)
        df2_monate_grouped = df2_monate_grouped.sort_values('Monat_Zahl')

        # Liniendiagramm erstellen
        fig_line = px.line(
            df2_monate_grouped,
            x='Monat',
            y='Gewicht',
            color='Jahr',
            title="Entwicklung des Gewichts über die Monate hinweg",
            labels={'Monat': 'Monat', 'Gewicht': 'Gesamtgewicht', 'Jahr': 'Jahr'}
        )

        # Größe des Liniendiagramms anpassen
        fig_line.update_layout(width=800, height=500)

        # Liniendiagramm anzeigen
        st.plotly_chart(fig_line)

    # Zeige den Data Preview für Tab 2 an
    with st.expander("Data Preview für Dashboard 2"):
        st.dataframe(df2)

# Tab 3: Dashboard 3 - Weitere Visualisierungen
with tab3:
    st.subheader("Dashboard 3 - Transporte_xlsx")

    # Datei automatisch von URL herunterladen
    file_dashboard3 = download_file_from_url(url_dashboard3)

    if file_dashboard3 is not None:
        # Checkbox zur Bestätigung durch den Benutzer
        if st.checkbox("Bestätigen Sie die Datei für Dashboard 3"):
            df3 = load_data(file_dashboard3)
            st.write(df3.head())  # Zeige einen kurzen Überblick über die Daten
            # Restliche Visualisierungen hier hinzufügen...
    else:
        st.error("Die Datei für Dashboard 3 konnte nicht heruntergeladen werden.")
        
    # Lade die Daten für Dashboard 3 (nur die Datei hochladen und keine weiteren Veränderungen vornehmen)
    @st.cache_data
    def load_data_tab3(file):
        # Excel-Datei laden
        data = pd.read_excel(file)
        
        # Entferne führende und nachfolgende Leerzeichen von allen Spaltennamen
        data.columns = data.columns.str.strip()

        # Konvertiere die Spalten "Fahrbeginn Zeit" und "Ende Zeit" in Zeitformat für die Berechnung der Zeitdifferenzen
        data['Fahrbeginn Zeit'] = pd.to_datetime(data['Fahrbeginn Zeit'], errors='coerce')
        data['Ende Zeit'] = pd.to_datetime(data['Ende Zeit'], errors='coerce')

        # Filtere Zeilen ohne gültige Zeitwerte
        data = data.dropna(subset=['Fahrbeginn Zeit', 'Ende Zeit'])

        return data

    # Verwende die Funktion, um die Datei für Tab 3 zu laden
    df3 = load_data_tab3(uploaded_file_3)

    # Extrahiere die ersten zwei Zeichen der Spalten "Quell-Platz" und "Ziel-Platz"
    df3['Quell-Bereich'] = df3['Quell-Platz'].str[:2]
    df3['Ziel-Bereich'] = df3['Ziel-Platz'].str[:2]

    # Filtere Ziel-Bereiche und Quell-Bereiche aus, die '00' enthalten
    df3 = df3[df3['Ziel-Bereich'] != '00']
    df3 = df3[df3['Quell-Bereich'] != '00']

    # Setze zwei Diagramme nebeneinander
    col1, col2 = st.columns(2)

    # Balkendiagramm für Ziel-Bereiche in col1
    with col1:
        # Zähle die Anzahl der Transporte pro Zielbereich
        df3_grouped = df3.groupby('Ziel-Bereich').size().reset_index(name='Anzahl Transporte')

        # Einzigartige Ziel-Bereiche erfassen
        unique_ziel_bereiche = df3_grouped['Ziel-Bereich'].unique()

        # Sortiere die Ziel-Bereiche nach der Bedingung WE -> numerisch -> WA
        sorted_ziel_bereiche = ['WE'] + sorted(
            [bereich for bereich in unique_ziel_bereiche if bereich not in ['WE', 'WA']],
            key=lambda x: int(x) if x.isdigit() else float('inf')
        ) + ['WA']

        # Wandle Ziel-Bereich in eine kategorische Spalte mit der festgelegten Reihenfolge
        df3_grouped['Ziel-Bereich'] = pd.Categorical(
            df3_grouped['Ziel-Bereich'], 
            categories=sorted_ziel_bereiche,
            ordered=True
        )

        # Füge "Bereich" vor die Zielbereiche hinzu, außer bei "WE" und "WA"
        df3_grouped['Ziel-Bereich'] = df3_grouped['Ziel-Bereich'].apply(lambda x: f"Bereich {x}" if x not in ['WE', 'WA'] else x)

        # Sortiere das DataFrame entsprechend der definierten Kategorie-Reihenfolge
        df3_grouped = df3_grouped.sort_values('Ziel-Bereich')

        # Erstelle ein Balkendiagramm für Ziel-Bereiche
        fig_balken = px.bar(
            df3_grouped,
            x='Ziel-Bereich',
            y='Anzahl Transporte',
            labels={'Ziel-Bereich': 'Ziel-Bereich', 'Anzahl Transporte': 'Anzahl der Transporte'},
            title="Anzahl der Transporte pro Zielbereich"
        )

        # Größe des Diagramms anpassen
        fig_balken.update_layout(width=800, height=500)

        # Zeige das Balkendiagramm an
        st.plotly_chart(fig_balken)

    # Balkendiagramm für Quell-Bereiche in col2
    with col2:
        # Zähle die Anzahl der Transporte pro Quellbereich
        df3_quell_grouped = df3.groupby('Quell-Bereich').size().reset_index(name='Anzahl Transporte')

        # Einzigartige Quell-Bereiche erfassen
        unique_quell_bereiche = df3_quell_grouped['Quell-Bereich'].unique()

        # Sortiere die Quell-Bereiche nach der Bedingung WE -> numerisch -> WA
        sorted_quell_bereiche = ['WE'] + sorted(
            [bereich for bereich in unique_quell_bereiche if bereich not in ['WE', 'WA']],
            key=lambda x: int(x) if x.isdigit() else float('inf')
        ) + ['WA']

        # Wandle Quell-Bereich in eine kategorische Spalte mit der festgelegten Reihenfolge
        df3_quell_grouped['Quell-Bereich'] = pd.Categorical(
            df3_quell_grouped['Quell-Bereich'], 
            categories=sorted_quell_bereiche,
            ordered=True
        )

        # Füge "Bereich" vor die Quellbereiche hinzu, außer bei "WE" und "WA"
        df3_quell_grouped['Quell-Bereich'] = df3_quell_grouped['Quell-Bereich'].apply(lambda x: f"Bereich {x}" if x not in ['WE', 'WA'] else x)

        # Sortiere das DataFrame entsprechend der definierten Kategorie-Reihenfolge
        df3_quell_grouped = df3_quell_grouped.sort_values('Quell-Bereich')

        # Erstelle ein Balkendiagramm für Quell-Bereiche
        fig_balken_quell = px.bar(
            df3_quell_grouped,
            x='Quell-Bereich',
            y='Anzahl Transporte',
            labels={'Quell-Bereich': 'Quell-Bereich', 'Anzahl Transporte': 'Anzahl der Transporte'},
            title="Anzahl der Transporte pro Quellbereich"
        )

        # Größe des Diagramms anpassen
        fig_balken_quell.update_layout(width=800, height=500)

        # Zeige das Balkendiagramm an
        st.plotly_chart(fig_balken_quell)

    # Eckdaten als Labels unter den Diagrammen
    st.subheader("Wichtige Kennzahlen")

    # Berechnung der Gesamtanzahl der Transporte
    gesamt_transporte = df3.shape[0]

    # Berechnung der Transportdauer in Minuten (Differenz zwischen "Fahrbeginn Zeit" und "Ende Zeit")
    df3['Transportdauer'] = (df3['Ende Zeit'] - df3['Fahrbeginn Zeit']).dt.total_seconds() / 60

    # Berechne die gesamte Transportdauer (Summe aller Transportdauern)
    gesamt_transportdauer = df3['Transportdauer'].sum()

    # Berechne die durchschnittliche Transportdauer (Gesamtdauer geteilt durch die Anzahl der Aufträge) und füge 4 Minuten hinzu
    durchschnitt_dauer = (gesamt_transporte/ gesamt_transportdauer)

    # Berechnung des Gesamtgewichts in Kilogramm
    gesamt_gewicht = df3['Gewicht'].sum()

    # Zeige die Kennzahlen als Labels an
    col3, col4, col5 = st.columns(3)

    col3.metric(label="Gesamtanzahl der Transporte", value=f"{gesamt_transporte}")
    col4.metric(label="Durchschnittliche Transportdauer", value=f"{durchschnitt_dauer:.2f} Minuten")
    col5.metric(label="Gesamtgewicht transportiert", value=f"{gesamt_gewicht:.2f} kg")

    # Zeige den Data Preview für Tab 3 an
    with st.expander("Data Preview für Dashboard 3"):
        st.dataframe(df3)

with tab4:
    st.subheader("Dashboard 4 - Speditionspreise_xlsx")

    # Datei automatisch von URL herunterladen
    file_dashboard4 = download_file_from_url(url_dashboard4)

    if file_dashboard4 is not None:
        # Checkbox zur Bestätigung durch den Benutzer
        if st.checkbox("Bestätigen Sie die Datei für Dashboard 4"):
            df4 = load_data(file_dashboard4)
            st.write(df4.head())  # Zeige einen kurzen Überblick über die Daten
            # Restliche Visualisierungen hier hinzufügen...
    else:
        st.error("Die Datei für Dashboard 4 konnte nicht heruntergeladen werden.")

    # Lade die Preisdaten ab der ersten Zeile (keine Zeilen überspringen)
    @st.cache_data
    def load_price_data(file):
        xls = pd.ExcelFile(file)
        df_spedition_1 = pd.read_excel(xls, sheet_name="Dachser")
        df_spedition_2 = pd.read_excel(xls, sheet_name="Schenker")
        df_spedition_3 = pd.read_excel(xls, sheet_name="Rhenus")
        df_spedition_4 = pd.read_excel(xls, sheet_name="Kühne")

        # PLZ-Spalte als String mit führenden Nullen
        df_spedition_1['PLZ'] = df_spedition_1['PLZ'].astype(str).str.zfill(2)
        df_spedition_2['PLZ'] = df_spedition_2['PLZ'].astype(str).str.zfill(2)
        df_spedition_3['PLZ'] = df_spedition_3['PLZ'].astype(str).str.zfill(2)
        df_spedition_4['PLZ'] = df_spedition_4['PLZ'].astype(str).str.zfill(2)

        return df_spedition_1, df_spedition_2, df_spedition_3, df_spedition_4

    # Daten laden
    df_spedition_1, df_spedition_2, df_spedition_3, df_spedition_4 = load_price_data(uploaded_file_4)

    # Entferne Eurozeichen und konvertiere Preise in numerische Werte
    df_spedition_1.iloc[:, 1:] = df_spedition_1.iloc[:, 1:].replace({'€': '', ',': '.'}, regex=True).astype(float)
    df_spedition_2.iloc[:, 1:] = df_spedition_2.iloc[:, 1:].replace({'€': '', ',': '.'}, regex=True).astype(float)
    df_spedition_3.iloc[:, 1:] = df_spedition_3.iloc[:, 1:].replace({'€': '', ',': '.'}, regex=True).astype(float)
    df_spedition_4.iloc[:, 1:] = df_spedition_4.iloc[:, 1:].replace({'€': '', ',': '.'}, regex=True).astype(float)

    # Gewichte zur Auswahl im Dropdown-Menü
    gewichte = df_spedition_1.columns[1:]  # Annahme: Gewichtsspalten beginnen ab der 2. Spalte
    
    # Gewichtsauswahl
    gewaehltes_gewicht = st.selectbox("Wähle ein Gewicht", options=gewichte)

    # Setze zwei Diagramme nebeneinander
    col1, col2 = st.columns(2)

    with col1:
        if 'PLZ' not in df_spedition_1.columns:
            st.error("Die Spalte 'PLZ' existiert nicht in den Daten. Überprüfe die Spaltennamen.")
        else:
            # Liniendiagramm für das gewählte Gewicht
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=df_spedition_1['PLZ'], y=df_spedition_1[gewaehltes_gewicht], mode='lines', name='Dachser', line=dict(color='yellow')))
            fig.add_trace(go.Scatter(x=df_spedition_2['PLZ'], y=df_spedition_2[gewaehltes_gewicht], mode='lines', name='Schenker', line=dict(color='red')))
            fig.add_trace(go.Scatter(x=df_spedition_3['PLZ'], y=df_spedition_3[gewaehltes_gewicht], mode='lines', name='Rhenus', line=dict(color='blue')))
            fig.add_trace(go.Scatter(x=df_spedition_4['PLZ'], y=df_spedition_4[gewaehltes_gewicht], mode='lines', name='Kühne', line=dict(color='green')))
            fig.update_layout(title=f"Preise für Gewicht {gewaehltes_gewicht} kg", xaxis_title="Postleitzahl", yaxis_title="Preis (€)", width=800, height=500)
            st.plotly_chart(fig)

    with col2:
        # Berechnung der Bestpreise pro Gewicht
        df_bestpreis = pd.DataFrame(df_spedition_1['PLZ'])
        for gewicht in gewichte:
            df_bestpreis[gewicht] = df_spedition_1[gewicht].combine(df_spedition_2[gewicht], min).combine(df_spedition_3[gewicht], min).combine(df_spedition_4[gewicht], min)

        # Liniendiagramm für Bestpreise bei dem ausgewählten Gewicht
        fig_bestpreis = go.Figure()
        fig_bestpreis.add_trace(go.Scatter(x=df_bestpreis['PLZ'], y=df_bestpreis[gewaehltes_gewicht], mode='lines', name='Bestpreis', line=dict(color='lightgreen')))
        fig_bestpreis.update_layout(title=f"Bestpreise für Gewicht {gewaehltes_gewicht} kg", xaxis_title="Postleitzahl", yaxis_title="Bestpreis (€)", width=800, height=500)
        st.plotly_chart(fig_bestpreis)

    # Button zum Exportieren der Bestpreisliste für alle Gewichte und PLZ in Excel
    if st.button("Bestpreisliste für alle Gewichte als Excel exportieren"):
        # Erstelle eine neue Excel-Datei im Speicher
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')

        # Füge die Bestpreisliste zur Excel-Datei hinzu
        df_bestpreis.to_excel(writer, index=False, sheet_name='Bestpreisliste')

        # Speichere den Writer
        writer.close()

        # Lade die erstellte Excel-Datei
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        sheet = workbook['Bestpreisliste']

        # Definiere Farben für die Speditionen
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        gray_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        # Färbe die Zellen der Bestpreisliste basierend auf der jeweiligen Spedition
        for row in range(2, sheet.max_row + 1):
            for col in range(2, sheet.max_column + 1):
                zelle_value = sheet.cell(row=row, column=col).value
                if zelle_value is None:
                    continue  # Überspringe leere Zellen
                # Check für Speditionen basierend auf den Daten für das jeweilige Gewicht
                if zelle_value in df_spedition_1.iloc[:, col-1].values:
                    sheet.cell(row=row, column=col).fill = yellow_fill
                elif zelle_value in df_spedition_2.iloc[:, col-1].values:
                    sheet.cell(row=row, column=col).fill = red_fill
                elif zelle_value in df_spedition_3.iloc[:, col-1].values:
                    sheet.cell(row=row, column=col).fill = blue_fill
                elif zelle_value in df_spedition_4.iloc[:, col-1].values:
                    sheet.cell(row=row, column=col).fill = green_fill
                else:
                    # Zellen grau einfärben, wenn keine Zuordnung möglich ist
                    sheet.cell(row=row, column=col).fill = gray_fill


        # Speichere die gefärbte Datei
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Biete die Datei zum Download an
        st.download_button(
            label="Download Bestpreisliste",
            data=output,
            file_name="Bestpreisliste.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Zeige den Data Preview für Tab 4 an
    with st.expander("Data Preview für Dashboard 4"):
        st.dataframe(df_spedition_1)

